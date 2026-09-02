#!/usr/bin/env python3
"""Buggy/non-buggy review sheet for George and Duncan — export and import.

    export   python scripts/export_buggy_review.py [--since 2025-01-01] [--out PATH]
    import   python scripts/export_buggy_review.py --import FILE.xlsx [--apply]
    backfill python scripts/export_buggy_review.py --backfill [--apply]

**Export.** One tab per athlete, one row per parkrun from --since, carrying the
evidence behind each run so the athlete corrects against numbers rather than
recalling dates. `Confirmed` is theirs to fill in (dropdown: With buggy /
Without buggy).

The sheet deliberately carries **no estimate**. An earlier version embedded an
unsupervised threshold rule; that approach was abandoned, because the course
confound is larger than the buggy signal — a hard course (+20-33%) cannot be
told from a buggy (+10-14%) per-run, and calibrating the rule so it never
flagged Raju (who has never used a buggy) pushed the threshold to +50%, at which
point it caught almost nothing. There is no trained model yet either, so an
Estimate column would present a guess from a discarded method as if it meant
something. What is left is the evidence: `Baseline` and `Excess`, whose real
signal is a **run of consecutive same-sign deviations**, not any single row.

**Import.** Reads a returned sheet and writes parkrun.run_modes three ways:
  1. answered  -> source='manual', is_buggy as stated. Trains the model.
  2. blank     -> source='default', is_buggy=FALSE, and listed in the output.
                  An unanswered row is an assumption, not a confirmation.
  3. every run outside the sheet (Raju included) -> source='default', FALSE.
Prints what it would do; writes only with --apply.

**Backfill.** Rule 3 on its own, so the ~681 runs outside the sheet can be
labelled before it comes back — they are non-buggy whatever the answers, and
writing them early makes the estimator's anti-join converge.

Writes go to the SOURCE OF TRUTH (~/.config/parkrun/parkrun_local.duckdb), never
to the deploy snapshot; override with PARKRUN_DB for testing.

Build-time only. Needs openpyxl, deliberately NOT in requirements.txt (same
convention as scripts/build_logo.py) — install it into the dev venv.
"""

from __future__ import annotations

import argparse
import os
from datetime import date, datetime
from pathlib import Path

import duckdb
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parent.parent

# The two athletes with a buggy dimension. Raju never uses one, so he is not
# exported — every run of his is non-buggy by construction.
ATHLETES = ("George", "Duncan")

# Baseline selection. `event` is the clean, course-controlled baseline; `window`
# is the fallback when there is too little history at that course.
BASE_MIN_EVENT_RUNS = 2
BASE_MIN_WINDOW_RUNS = 8
BASE_WINDOW_DAYS = 182

WITH_BUGGY = "With buggy"
WITHOUT_BUGGY = "Without buggy"

COLUMNS = [
    ("Date", 12),
    ("parkrun", 38),
    ("Time", 9),
    ("Position", 9),
    ("Age grade", 11),
    ("Baseline", 10),
    ("Excess", 9),
    ("Confirmed", 14),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")


SNAPSHOT = REPO / "data" / "parkrun_snapshot.duckdb"


def resolve_db(for_write: bool = False) -> str:
    """Same fallback order as app.py: env var, local source of truth, snapshot.

    A write must never land on the committed deploy snapshot — it is a build
    artefact, regenerated from the source of truth on every refresh, so a label
    written there would be silently destroyed.
    """
    env = os.environ.get("PARKRUN_DB")
    db = os.path.expanduser(env) if env else None
    if db is None:
        local = Path.home() / ".config" / "parkrun" / "parkrun_local.duckdb"
        db = str(local) if local.exists() else str(SNAPSHOT)
    if for_write and Path(db).resolve() == SNAPSHOT.resolve():
        raise SystemExit(
            f"refusing to write labels to the deploy snapshot ({SNAPSHOT}).\n"
            f"It is rebuilt from the source of truth on every refresh, so the "
            f"labels would be destroyed. Point at "
            f"~/.config/parkrun/parkrun_local.duckdb (or set PARKRUN_DB)."
        )
    return db


def load_results(db: str) -> pd.DataFrame:
    con = duckdb.connect(db, read_only=True)
    try:
        df = con.execute(
            """
            SELECT r.athlete_id, a.athlete_name, r.run_date, r.event_id,
                   e.short_name, r.time, r.time_seconds, r.position, r.age_grade
            FROM parkrun.results r
            JOIN parkrun.athletes a USING (athlete_id)
            JOIN parkrun.events e USING (event_id)
            """
        ).fetchdf()
    finally:
        con.close()
    df["run_date"] = pd.to_datetime(df["run_date"])
    return df


def fmt_time(seconds: float) -> str:
    s = int(round(seconds))
    return f"{s // 60}:{s % 60:02d}"


def add_baselines(df: pd.DataFrame) -> pd.DataFrame:
    """Attach a comparison baseline and the excess over it to every run.

    This is evidence for a human, not a verdict. Baselines, first rule that
    fires:
      1. `event`  — median of that athlete's OTHER runs at the same parkrun.
                    Course-controlled, so the only clean baseline.
      2. `window` — 25th percentile of their runs within +/-182 days. q25 not
                    median: a buggy only ever makes you slower, so the fast tail
                    is the buggy-immune part of the distribution. Note this is
                    biased high by construction — it is a rough guide, and it is
                    why no threshold is drawn on it.
      3. none     — too little history either way; Baseline and Excess are left
                    blank rather than guessed.
    """
    out = []
    for i, r in df.iterrows():
        others = df[(df.athlete_id == r.athlete_id) & (df.index != i)]
        same_event = others[others.event_id == r.event_id]
        if len(same_event) >= BASE_MIN_EVENT_RUNS:
            out.append((same_event.time_seconds.median(), "event", len(same_event)))
            continue
        near = others[
            (others.run_date - r.run_date).abs().dt.days <= BASE_WINDOW_DAYS
        ]
        if len(near) >= BASE_MIN_WINDOW_RUNS:
            out.append((near.time_seconds.quantile(0.25), "window", len(near)))
        else:
            out.append((np.nan, "none", 0))

    df[["expected", "basis", "n_base"]] = pd.DataFrame(out, index=df.index)
    df["excess"] = df.time_seconds / df.expected - 1
    return df


def write_sheet(ws, rows: pd.DataFrame) -> None:
    ws.append([c for c, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for _, r in rows.iterrows():
        ws.append(
            [
                r.run_date.date(),
                r.short_name,
                r.time,
                int(r.position) if pd.notna(r.position) else None,
                r.age_grade,
                fmt_time(r.expected) if pd.notna(r.expected) else "",
                round(float(r.excess), 4) if pd.notna(r.excess) else None,
                None,
            ]
        )

    for idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    excess_idx = [c for c, _ in COLUMNS].index("Excess")
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(COLUMNS)):
        row[0].number_format = "yyyy-mm-dd"
        row[excess_idx].number_format = "+0.0%;-0.0%"

    # showDropDown is inverted in the spec: False means "show the arrow".
    # showErrorMessage must be True or Google Sheets discards the validation
    # entirely on import.
    dv = DataValidation(
        type="list",
        formula1=f'"{WITH_BUGGY},{WITHOUT_BUGGY}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Pick one",
        error=f"Choose {WITH_BUGGY} or {WITHOUT_BUGGY}.",
    )
    ws.add_data_validation(dv)
    confirmed_col = get_column_letter(len(COLUMNS))
    dv.add(f"{confirmed_col}2:{confirmed_col}{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def read_sheet(path: Path) -> pd.DataFrame:
    """Read a returned review workbook into (athlete_name, run_date, parkrun,
    confirmed) rows. One tab per athlete; unknown tabs are ignored."""
    book = load_workbook(path, data_only=True)
    rows = []
    for name in ATHLETES:
        if name not in book.sheetnames:
            print(f"  WARNING: no '{name}' tab in {path.name} — skipped")
            continue
        ws = book[name]
        header = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(header)}
        for missing in ("Date", "parkrun", "Confirmed"):
            if missing not in idx:
                raise SystemExit(f"'{name}' tab has no {missing!r} column")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["Date"]] is None:
                continue
            d = row[idx["Date"]]
            rows.append({
                "athlete_name": name,
                "run_date": pd.Timestamp(d).date(),
                "short_name": row[idx["parkrun"]],
                "confirmed": (row[idx["Confirmed"]] or "").strip(),
            })
    return pd.DataFrame(rows)


def _label_frame(con, sheet: pd.DataFrame | None) -> pd.DataFrame:
    """Every result, with the label that should be written for it.

    Three ways, per the plan:
      1. answered on the sheet -> manual, is_buggy as stated
      2. blank on the sheet    -> default, FALSE (an assumption, not a
                                  confirmation — reported so it can be checked)
      3. not on the sheet      -> default, FALSE
    """
    runs = con.execute(
        """
        SELECT r.athlete_id, a.athlete_name, r.run_date, r.event_id, e.short_name
        FROM parkrun.results r
        JOIN parkrun.athletes a USING (athlete_id)
        JOIN parkrun.events e USING (event_id)
        """
    ).fetchdf()
    runs["run_date"] = pd.to_datetime(runs["run_date"]).dt.date
    runs["is_buggy"] = False
    runs["source"] = "default"
    runs["reason"] = "not reviewed — assumed non-buggy"

    if sheet is None or sheet.empty:
        return runs

    # Match on (athlete, date, parkrun). Date alone is not a key: an athlete can
    # run two events on the same day (decision 7), which is exactly why results
    # is keyed on the event too.
    key = ["athlete_name", "run_date", "short_name"]
    merged = runs.merge(sheet, on=key, how="left", indicator=True)
    unmatched = sheet.merge(runs, on=key, how="left", indicator=True)
    orphans = unmatched[unmatched["_merge"] == "left_only"]
    if not orphans.empty:
        print(f"  WARNING: {len(orphans)} sheet row(s) match no run — ignored:")
        for _, o in orphans.iterrows():
            print(f"    {o.athlete_name} {o.run_date} {o.short_name}")

    on_sheet = merged["_merge"] == "both"
    answered = on_sheet & merged["confirmed"].isin([WITH_BUGGY, WITHOUT_BUGGY])
    merged.loc[answered, "source"] = "manual"
    merged.loc[answered, "is_buggy"] = merged.loc[answered, "confirmed"] == WITH_BUGGY
    merged.loc[answered, "reason"] = "confirmed by the athlete"
    merged["blank_on_sheet"] = on_sheet & ~answered
    bad = on_sheet & ~answered & merged["confirmed"].fillna("").ne("")
    if bad.any():
        print(f"  WARNING: {int(bad.sum())} row(s) have an unrecognised "
              f"Confirmed value — treated as blank:")
        for v in sorted(merged.loc[bad, "confirmed"].unique()):
            print(f"    {v!r}")
    return merged


def write_labels(rows: pd.DataFrame, con, apply: bool) -> None:
    """Insert the labels. Never overwrites an existing row: a hand correction
    outranks this import permanently, and re-running must not undo one."""
    existing = con.execute("SELECT count(*) FROM parkrun.run_modes").fetchone()[0]
    stage = rows[["athlete_id", "run_date", "event_id", "is_buggy", "source",
                  "reason"]]
    con.register("label_stage", stage)
    new = con.execute(
        """
        SELECT count(*) FROM label_stage s
        WHERE NOT EXISTS (SELECT 1 FROM parkrun.run_modes m
                          WHERE m.athlete_id = s.athlete_id
                            AND m.run_date  = s.run_date
                            AND m.event_id  = s.event_id)
        """
    ).fetchone()[0]
    print(f"\n  run_modes holds {existing} row(s); this would insert {new} new "
          f"and leave {len(stage) - new} existing row(s) untouched")
    if not apply:
        print("  DRY RUN — nothing written. Re-run with --apply to write.")
        con.unregister("label_stage")
        return
    con.execute(
        """
        INSERT INTO parkrun.run_modes
              (athlete_id, run_date, event_id, is_buggy, source, confidence,
               reason, set_at)
        SELECT s.athlete_id, s.run_date, s.event_id, s.is_buggy, s.source,
               NULL, s.reason, now()
        FROM label_stage s
        WHERE NOT EXISTS (SELECT 1 FROM parkrun.run_modes m
                          WHERE m.athlete_id = s.athlete_id
                            AND m.run_date  = s.run_date
                            AND m.event_id  = s.event_id)
        """
    )
    con.unregister("label_stage")
    print(f"  WROTE {new} row(s); run_modes now holds "
          f"{con.execute('SELECT count(*) FROM parkrun.run_modes').fetchone()[0]}")


def report(rows: pd.DataFrame) -> None:
    """Class counts per athlete. Phase 4 gates on 8 labels per class, so this is
    the number that says whether an athlete can be modelled at all."""
    print("\n  per athlete:")
    for name, g in rows.groupby("athlete_name"):
        man = g[g.source == "manual"]
        buggy = int(man.is_buggy.sum())
        nonbuggy = int((~man.is_buggy).sum())
        gate = "" if min(buggy, nonbuggy) >= 8 else "   <- below the 8-per-class gate"
        print(f"    {name:<8} {len(g):>4} runs  "
              f"manual: {buggy} buggy / {nonbuggy} non-buggy  "
              f"default: {len(g) - len(man)}{gate}")
    if "blank_on_sheet" in rows and rows.blank_on_sheet.any():
        blanks = rows[rows.blank_on_sheet]
        print(f"\n  {len(blanks)} row(s) left BLANK on the sheet — assumed "
              f"non-buggy, not used for training. Check these:")
        for _, b in blanks.sort_values(["athlete_name", "run_date"]).iterrows():
            print(f"    {b.athlete_name:<8} {b.run_date}  {b.short_name}")


def do_import(path: Path | None, apply: bool) -> None:
    db = resolve_db(for_write=True)
    sheet = read_sheet(path) if path else None
    if sheet is not None:
        print(f"  read {len(sheet)} row(s) from {path.name}")
    con = duckdb.connect(db)
    try:
        rows = _label_frame(con, sheet)
        report(rows)
        write_labels(rows, con, apply)
    finally:
        con.close()
    print(f"\n  target DB : {db}")


def export(args) -> None:
    db = resolve_db()
    df = add_baselines(load_results(db))

    out = Path(args.out) if args.out else (
        REPO / "data" / f"buggy_review_{date.today():%Y-%m-%d}.xlsx"
    )

    # Built directly with openpyxl rather than through pandas' ExcelWriter: the
    # rows need per-cell formatting and a dropdown, and the workbook needs the
    # Google Sheets fixes below, neither of which to_excel can express.
    book = Workbook()
    book.remove(book.active)
    # openpyxl emits an empty <workbookProtection/> by default. Google Sheets
    # refuses to import a workbook carrying it, so drop it.
    book.security = None

    for name in ATHLETES:
        rows = df[
            (df.athlete_name == name) & (df.run_date >= args.since)
        ].sort_values("run_date").copy()
        ws = book.create_sheet(title=name)
        write_sheet(ws, rows)
        print(f"{name:8} {len(rows):3} runs since {args.since}")

    book.save(out)

    print(f"\nsource DB : {db}")
    print(f"written   : {out}")
    print(f"generated : {datetime.now():%Y-%m-%d %H:%M}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--since", default="2025-01-01")
    ap.add_argument("--out", default=None)
    ap.add_argument("--import", dest="import_path", type=Path, default=None,
                    help="import a returned review workbook")
    ap.add_argument("--backfill", action="store_true",
                    help="label every run 'default' non-buggy, no sheet needed")
    ap.add_argument("--apply", action="store_true",
                    help="actually write (import/backfill print a dry run "
                         "otherwise)")
    args = ap.parse_args()

    if args.import_path and args.backfill:
        raise SystemExit("--import and --backfill are alternatives")
    if args.import_path:
        do_import(args.import_path, args.apply)
    elif args.backfill:
        do_import(None, args.apply)
    else:
        export(args)


if __name__ == "__main__":
    main()
